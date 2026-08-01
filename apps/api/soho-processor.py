"""
Excel transformation: Data-input.xlsx  →  Form-output.xlsx (Daily Production Report)
Formatting matches the sample output 100%: fills, fonts, merges, row heights, column widths.
"""
import argparse
import io
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Color constants ────────────────────────────────────────────────────────────
C_HEADER_GREEN  = 'FF99FFCC'   # row-2 header background
C_COL_C         = 'FFFBE2D5'   # col C (FACTORY) & color-total rows  #FBE2D5
C_FORM_SUBTOTAL = 'FFF2CEEF'   # form/group subtotal (REGULAR-xxx, TALL-xxx)  #F2CEEF
C_PO_SUBTOTAL   = 'FFDAF2D0'   # PO-region subtotal (Total PO XXX-ECOM …)  #DAF2D0
C_GRAND_TOTAL   = 'FFCC99FF'   # grand TOTAL row
C_FONT_BLUE     = 'FF0070C0'   # metadata cell text
C_FONT_DARK     = 'FF0C0C0C'
C_FONT_RED      = 'FFFF0000'   # K, M, O header columns

# ── Column widths ──────────────────────────────────────────────────────────────
COL_WIDTHS = {
    'A': 8,    'B': 15.78, 'C': 24.0,  'D': 17.78,
    'E': 12.22,'F': 15.66, 'G': 11.44, 'H': 9.78,
    'I': 15.78,'J': 14.66, 'K': 15.44, 'L': 15.44,
    'M': 15.44,'N': 15.44, 'O': 15.44, 'P': 15.44,
}

DETAIL_HEADERS = [
    'LINE No','STYLE','FACTORY','PO','GROUP','COLOR','INSEAM','SIZE',
    'QTY (order)','CUTTING QTY','IN PUT/DAY','TOTAL IN PUT',
    'OUTPUT/ DAY','TOTAL OUTPUT','Warehousing','Put in carton',
]
SUMMARY_HEADERS = [
    None,'STYLE','FACTORY','PO',None,'COLOR','INSEAM','SIZE',
    'QTY (order)','CUTTING QTY','IN PUT/DAY','TOTAL IN PUT',
    'OUTPUT/ DAY','TOTAL OUTPUT','Warehousing','Put in carton',
]

# Column letters I-P (cols 9-16) for formula generation
_FORMULA_COLS = list('IJKLMNOP')

# ── Style helpers ──────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def _font(bold=True, size=11, color=C_FONT_DARK):
    return Font(bold=bold, size=size, color=color)

def _align(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def _apply(cell, fill=None, font=None, align=None, border=None):
    if fill:   cell.fill      = fill
    if font:   cell.font      = font
    if align:  cell.alignment = align
    if border: cell.border    = border

# ── Formula builders ──────────────────────────────────────────────────────────
def _f_range(col, start, end):
    """=SUM(X{start}:X{end})"""
    return f'=SUM({col}{start}:{col}{end})'

def _f_add(col, rows):
    """=+X{r1}+X{r2}+..."""
    return '=+' + '+'.join(f'{col}{r}' for r in rows)

def _f_sum(col, rows):
    """=+SUM(X{r1},X{r2},...)"""
    return f'=+SUM({",".join(f"{col}{r}" for r in rows)})'

# ── Input parser ──────────────────────────────────────────────────────────────
def _pad(row, n=35):
    r = list(row)
    while len(r) < n:
        r.append(None)
    return r

def _detect_style_groups(groups_row, sizes_row):
    """Scan up to col 24 for style group headers, stop at 'TOTAL'."""
    style_groups = []
    current = None
    for col in range(1, 25):
        g = groups_row[col] if col < len(groups_row) else None
        s = sizes_row[col]  if col < len(sizes_row)  else None
        g_str = str(g).strip() if g else ''
        if g_str.upper() == 'TOTAL':
            break
        if g_str and g_str.upper() not in ('', 'PACK'):
            current = {'name': g_str, 'sizes': []}
            style_groups.append(current)
        if current and s and str(s).strip() and str(s).strip().upper() != 'TOTAL':
            current['sizes'].append((col, str(s).strip()))
    return style_groups

def _has_numeric_qtys(row, style_groups):
    """True if at least one style group column has a numeric quantity."""
    for sg in style_groups:
        for col_idx, _ in sg['sizes']:
            v = row[col_idx] if col_idx < len(row) else None
            if isinstance(v, (int, float)):
                return True
    return False

def _parse_block(rows, start_idx, po_col, region_col):
    groups_row   = _pad(rows[start_idx + 1] if start_idx + 1 < len(rows) else [])
    sizes_row    = _pad(rows[start_idx + 2] if start_idx + 2 < len(rows) else [])
    style_groups = _detect_style_groups(groups_row, sizes_row)

    colors   = []
    po_num   = None
    region   = None

    ci = start_idx + 3
    while ci < len(rows):
        row  = _pad(rows[ci])
        col0 = str(row[0]).strip() if row[0] is not None else ''

        # Block terminators
        if col0.upper() in ('TOTAL', 'SUMMARY') or col0 == '合计':
            ci += 1
            break
        if not col0:
            ci += 1
            continue
        # Skip formula/subtotal rows that have no numeric quantities
        if not _has_numeric_qtys(row, style_groups):
            ci += 1
            continue

        color_name = col0
        if po_num is None:
            raw = row[po_col] if po_col < len(row) else None
            if raw is not None and not (isinstance(raw, str) and raw.startswith('=')):
                po_num = str(int(raw)) if isinstance(raw, float) and raw == int(raw) else str(raw)
        if region is None:
            raw = row[region_col] if region_col < len(row) else None
            if raw is not None and not (isinstance(raw, str) and raw.startswith('=')):
                region = str(raw).strip()

        group_qtys = {}
        for sg in style_groups:
            qtys = {}
            for col_idx, size_label in sg['sizes']:
                v = row[col_idx] if col_idx < len(row) else None
                qtys[size_label] = int(v) if isinstance(v, (int, float)) and v else 0
            group_qtys[sg['name']] = qtys
        colors.append({'color': color_name, 'group_qtys': group_qtys})
        ci += 1

    po_str     = po_num or 'TBA'
    region_str = region or ''
    po_name    = f"{po_str}-{region_str}" if region_str else po_str

    return {'po_name': po_name, 'style_groups': style_groups, 'colors': colors}, ci

def _parse_input(wb):
    ws   = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    # Find style# dynamically: locate 'STYLE#' label, take first non-None value after it
    style_num = ''
    for row_idx in range(min(8, len(rows))):
        r = _pad(rows[row_idx])
        for ci, val in enumerate(r):
            if val and str(val).strip().upper() == 'STYLE#':
                for ni in range(ci + 1, len(r)):
                    if r[ni] is not None:
                        style_num = str(r[ni]).strip()
                        break
                break
        if style_num:
            break

    # Default PO# / REGION column indices (overridden per block from header row)
    po_col     = 17
    region_col = 18

    po_blocks = []
    i = 0
    while i < len(rows):
        row  = _pad(rows[i])
        col0 = str(row[0]).strip() if row[0] is not None else ''
        col1 = str(row[1]).strip() if row[1] is not None else ''
        if col0.upper() == 'COLOR' and 'SIZE' in col1.upper():
            # Detect PO# and REGION column positions from this header row
            for ci, val in enumerate(row):
                v = str(val).strip().upper() if val else ''
                if v == 'PO#':
                    po_col = ci
                elif v == 'REGION':
                    region_col = ci
            block, i = _parse_block(rows, i, po_col, region_col)
            if block['colors']:
                po_blocks.append(block)
        else:
            i += 1
    if not po_blocks:
        raise ValueError('File bạn vừa chọn không đúng định dạng, vui lòng kiểm tra lại')
    return style_num, po_blocks

# ── Output writers ─────────────────────────────────────────────────────────────
def _write_title_row(ws, current_row):
    ws.append(['DAILY PRODUCTION REPORT'] + [None] * 15)
    ws.row_dimensions[current_row].height = 28.5
    ws.merge_cells(f'A{current_row}:P{current_row}')
    cell = ws.cell(row=current_row, column=1)
    _apply(cell,
           font=Font(bold=True, size=18, color=C_FONT_DARK),
           align=_align('center', 'center'),
           border=_thin_border())
    return current_row + 1

def _write_header_row(ws, current_row, headers):
    ws.append(headers)
    ws.row_dimensions[current_row].height = 27.6
    fill     = _fill(C_HEADER_GREEN)
    tb       = _thin_border()
    red_cols = {11, 13, 15}
    for c, val in enumerate(headers, 1):
        if val is None:
            continue
        cell    = ws.cell(row=current_row, column=c)
        h_align = 'right' if c >= 10 else 'center'
        fc      = C_FONT_RED if c in red_cols else C_FONT_DARK
        _apply(cell,
               fill=fill,
               font=Font(bold=True, size=11, color=fc),
               align=_align(h_align, 'center', wrap=True),
               border=tb)
    return current_row + 1

def _write_color_block(ws, current_row, style_num, factory, po_name,
                       group_name, color_name, qtys):
    """
    Write size rows for one color with vertical merges A-G.
    Returns (next_row, block_start, block_end).
    """
    size_labels = list(qtys.keys())
    n_sizes     = len(size_labels)
    block_start = current_row
    block_end   = current_row + n_sizes - 1
    tb          = _thin_border()

    col_c_fill = _fill(C_COL_C)
    blue_fnt   = Font(bold=True, size=11, color=C_FONT_BLUE)
    dark_fnt   = _font(bold=True)
    ctr        = _align('center', 'center', wrap=True)
    ctr_nw     = _align('center', 'center')

    for i, sz in enumerate(size_labels):
        r = current_row + i
        ws.row_dimensions[r].height = 18.75
        row_data = [None] * 16
        if i == 0:
            row_data[1] = style_num
            row_data[2] = factory
            row_data[3] = po_name
            row_data[4] = group_name
            row_data[5] = color_name
            row_data[6] = None   # INSEAM – blank per spec
        row_data[7] = sz
        row_data[8] = qtys[sz]
        ws.append(row_data)

    for r in range(block_start, block_end + 1):
        for c in range(1, 17):
            cell = ws.cell(row=r, column=c)
            cell.border = tb
            # Layer 1: all cells default white (no fill set = white)
            # Layer 2: col C gets C_COL_C
            if c == 3:
                cell.fill      = col_c_fill
                cell.font      = blue_fnt
                cell.alignment = ctr
            elif c in (2, 4, 5, 6, 7):   # B,D-G – blue font, white background
                cell.font      = blue_fnt
                cell.alignment = ctr
            elif c == 8:                  # H – SIZE
                cell.font      = dark_fnt
                cell.alignment = ctr
            elif c == 9:                  # I – QTY
                cell.font      = dark_fnt
                cell.alignment = ctr_nw

    # Merge A-G vertically across all size rows
    for col in range(1, 8):
        if n_sizes > 1:
            cl = get_column_letter(col)
            ws.merge_cells(f'{cl}{block_start}:{cl}{block_end}')
        ws.cell(row=block_start, column=col).alignment = ctr

    return block_end + 1, block_start, block_end

def _write_color_total(ws, current_row, block_start, block_end):
    """
    'Total ' row after each color block. Uses SUM formulas for I-P.
    Returns next_row (color_total_row = current_row).
    """
    ws.row_dimensions[current_row].height = 18.75
    row_data = ['Total '] + [None] * 6 + ['Total']
    for col in _FORMULA_COLS:
        row_data.append(_f_range(col, block_start, block_end))
    ws.append(row_data)

    # Layer 3: color total row → C_COL_C (#FBE2D5)
    fill    = _fill(C_COL_C)
    bold    = _font(bold=True)
    ctr     = _align('center', 'center')
    tb      = _thin_border()
    num_fmt = '_(* #,##0_);_(* (#,##0);_(* -??_);_(@_)'

    for c in range(1, 17):
        cell = ws.cell(row=current_row, column=c)
        cell.fill      = fill
        cell.font      = bold
        cell.border    = tb
        cell.alignment = ctr
        if c >= 9:
            cell.number_format = num_fmt

    ws.merge_cells(f'A{current_row}:C{current_row}')
    return current_row + 1

def _write_group_subtotal(ws, current_row, po_name, group_name, color_total_rows):
    """
    'Total PO … - GROUP' row. Formulas sum all color_total_rows.
    Returns next_row (group_subtotal_row = current_row).
    """
    ws.row_dimensions[current_row].height = 27.0
    group_label = group_name.replace(' - ', ' ')
    label       = f'Total PO {po_name} - {group_label}'
    row_data    = [label] + [None] * 7
    for col in _FORMULA_COLS:
        row_data.append(_f_add(col, color_total_rows))
    ws.append(row_data)

    # Layer 4: form/group subtotal → C_FORM_SUBTOTAL (#F2CEEF)
    fill    = _fill(C_FORM_SUBTOTAL)
    bold    = _font(bold=True)
    ctr     = _align('center', 'center', wrap=True)
    tb      = _thin_border()
    num_fmt = '_(* #,##0_);_(* (#,##0);_(* -??_);_(@_)'

    for c in range(1, 17):
        cell = ws.cell(row=current_row, column=c)
        cell.fill      = fill
        cell.font      = bold
        cell.border    = tb
        cell.alignment = ctr
        if c >= 9:
            cell.number_format = num_fmt

    ws.merge_cells(f'A{current_row}:H{current_row}')
    return current_row + 1

def _write_po_total(ws, current_row, po_name, group_subtotal_rows):
    """
    'Total PO …' row for multi-group POs. Formulas sum all group_subtotal_rows.
    Returns next_row (po_total_row = current_row).
    """
    ws.row_dimensions[current_row].height = 27.0
    row_data = [f'Total PO {po_name}'] + [None] * 7
    for col in _FORMULA_COLS:
        row_data.append(_f_sum(col, group_subtotal_rows))
    ws.append(row_data)

    # Layer 5: PO-region subtotal → C_PO_SUBTOTAL (#DAF2D0)
    fill    = _fill(C_PO_SUBTOTAL)
    bold    = _font(bold=True)
    ctr     = _align('center', 'center', wrap=True)
    tb      = _thin_border()
    num_fmt = '_(* #,##0_);_(* (#,##0);_(* -??_);_(@_)'

    for c in range(1, 17):
        cell = ws.cell(row=current_row, column=c)
        cell.fill      = fill
        cell.font      = bold
        cell.border    = tb
        cell.alignment = ctr
        if c >= 9:
            cell.number_format = num_fmt

    ws.merge_cells(f'A{current_row}:H{current_row}')
    return current_row + 1

def _write_grand_total(ws, current_row, all_group_subtotal_rows):
    """Grand TOTAL row summing all group subtotals across every PO."""
    ws.row_dimensions[current_row].height = 30.0
    row_data = ['TOTAL'] + [None] * 7
    for col in _FORMULA_COLS:
        row_data.append(_f_sum(col, all_group_subtotal_rows))
    ws.append(row_data)

    fill    = _fill(C_GRAND_TOTAL)
    bold    = _font(bold=True, size=12)
    ctr     = _align('center', 'center', wrap=True)
    tb      = _thin_border()
    num_fmt = '_(* #,##0_);_(* (#,##0);_(* -??_);_(@_)'

    for c in range(1, 17):
        cell = ws.cell(row=current_row, column=c)
        cell.fill      = fill
        cell.font      = bold
        cell.border    = tb
        cell.alignment = ctr
        if c >= 9:
            cell.number_format = num_fmt

    ws.merge_cells(f'A{current_row}:H{current_row}')
    return current_row + 1

# ── Summary section ────────────────────────────────────────────────────────────
def _write_summary_section(ws, current_row, style_num, factory, po_summary):
    # Blank separator
    ws.row_dimensions[current_row].height = 9
    ws.append([None] * 16)
    current_row += 1

    current_row = _write_header_row(ws, current_row, SUMMARY_HEADERS)

    col_c_fill = _fill(C_COL_C)
    sub_fill   = _fill(C_FORM_SUBTOTAL)
    bold       = _font(bold=True)
    blue_fnt   = Font(bold=True, size=11, color=C_FONT_BLUE)
    ctr        = _align('center', 'center', wrap=True)
    tb         = _thin_border()

    all_summary_group_rows = []

    for po in po_summary:
        for group_name, group_total, all_colors in po['groups']:
            block_start = current_row
            block_end   = current_row + len(all_colors) - 1

            for i, (color_name, qty) in enumerate(all_colors):
                r = current_row + i
                ws.row_dimensions[r].height = 20.25
                ws.append([
                    None, style_num, factory, po['po_name'],
                    group_name, color_name, None, None, qty,
                ] + [None] * 7)

                for c in range(1, 17):
                    cell = ws.cell(row=r, column=c)
                    cell.border    = tb
                    cell.font      = blue_fnt if c in (2, 3, 4, 5, 6) else bold
                    cell.alignment = ctr
                    if c == 3:
                        cell.fill = col_c_fill

            # Merge B-E vertically for metadata
            n = len(all_colors)
            if n > 1:
                for col in (2, 3, 4, 5):
                    cl = get_column_letter(col)
                    ws.merge_cells(f'{cl}{block_start}:{cl}{block_end}')
                    ws.cell(row=block_start, column=col).alignment = ctr

            # Group total row for summary
            tr = current_row + n
            ws.row_dimensions[tr].height = 20.25
            ws.append([None, None, 'Total'] + [None] * 5 + [group_total] + [None] * 7)

            for c in range(1, 17):
                cell = ws.cell(row=tr, column=c)
                cell.fill      = sub_fill
                cell.font      = bold
                cell.border    = tb
                cell.alignment = ctr

            ws.merge_cells(f'C{tr}:H{tr}')
            all_summary_group_rows.append(tr)
            current_row = tr + 1

    # Summary grand total
    ws.row_dimensions[current_row].height = 20.25
    ws.append([None, 'TOTAL'] + [None] * 6 + [sum(
        po['groups'][gi][1] for po in po_summary for gi in range(len(po['groups']))
    )] + [None] * 7)
    for c in range(1, 17):
        cell = ws.cell(row=current_row, column=c)
        cell.fill      = _fill(C_GRAND_TOTAL)
        cell.font      = _font(bold=True, size=12)
        cell.border    = tb
        cell.alignment = ctr
    ws.merge_cells(f'B{current_row}:H{current_row}')

# ── Main entry point ───────────────────────────────────────────────────────────
def convert_excel(input_bytes: bytes, factory_name: str) -> bytes:
    import openpyxl
    wb_in = openpyxl.load_workbook(io.BytesIO(input_bytes))
    style_num, po_blocks = _parse_input(wb_in)

    wb_out = Workbook()
    ws     = wb_out.active
    ws.title = 'Output'

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Section 1: Detailed ────────────────────────────────────────────────
    current_row = 1
    current_row = _write_title_row(ws, current_row)
    current_row = _write_header_row(ws, current_row, DETAIL_HEADERS)

    all_group_subtotal_rows = []  # tracks every group-subtotal row for grand total
    po_summary = []

    for block in po_blocks:
        po_name      = block['po_name']
        active_groups = []
        group_subtotal_rows_this_po = []

        for sg in block['style_groups']:
            group_name    = sg['name']
            active_colors = [
                (c['color'], c['group_qtys'].get(group_name, {}))
                for c in block['colors']
                if any(v > 0 for v in c['group_qtys'].get(group_name, {}).values())
            ]
            if not active_colors:
                continue

            color_total_rows = []
            group_total      = 0
            color_totals     = {}

            for color_name, qtys in active_colors:
                color_sum = sum(qtys.values())
                current_row, block_start, block_end = _write_color_block(
                    ws, current_row, style_num, factory_name,
                    po_name, group_name, color_name, qtys)

                total_row   = current_row  # color total goes here
                current_row = _write_color_total(ws, current_row, block_start, block_end)
                color_total_rows.append(total_row)
                group_total += color_sum
                color_totals[color_name.strip().upper()] = color_sum

            group_subtotal_row = current_row
            current_row = _write_group_subtotal(
                ws, current_row, po_name, group_name, color_total_rows)

            all_group_subtotal_rows.append(group_subtotal_row)
            group_subtotal_rows_this_po.append(group_subtotal_row)

            all_colors_summary = [
                (c['color'], color_totals.get(c['color'].strip().upper()))
                for c in block['colors']
            ]
            active_groups.append((group_name, group_total, all_colors_summary))

        if len(group_subtotal_rows_this_po) > 1:
            current_row = _write_po_total(
                ws, current_row, po_name, group_subtotal_rows_this_po)

        po_summary.append({'po_name': po_name, 'groups': active_groups})

    current_row = _write_grand_total(ws, current_row, all_group_subtotal_rows)

    # ── Section 2: Summary ─────────────────────────────────────────────────
    _write_summary_section(ws, current_row, style_num, factory_name, po_summary)

    out = io.BytesIO()
    wb_out.save(out)
    out.seek(0)
    return out.read()


def main() -> int:
    parser = argparse.ArgumentParser(description='Convert SOHO Excel into Daily Production Report')
    parser.add_argument('--factory-name', required=True, help='Factory name to print in the output file')
    args = parser.parse_args()

    input_bytes = sys.stdin.buffer.read()
    if not input_bytes:
        print('Không có dữ liệu đầu vào', file=sys.stderr)
        return 1

    try:
        output_bytes = convert_excel(input_bytes, args.factory_name)
        sys.stdout.buffer.write(output_bytes)
        return 0
    except Exception as exc:
        print(f'Lỗi chuyển đổi: {exc}', file=sys.stderr)
        return 2


if __name__ == '__main__':
    raise SystemExit(main())
